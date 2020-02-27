from openpyxl import load_workbook
from practical_01.data_package.file_path_package import GetFilePath
class ExcelProcess:

    def write_back(self,sheetName,row,result=None,test_result=None):
        filePath=GetFilePath().get_test_data_file_path()
        wb = load_workbook(filePath)
        sheet=wb[sheetName]
        sheet.cell(row, 7).value = test_result
        sheet.cell(row,8).value=result
        wb.save(filePath)
        return sheet.cell(row,8).value


# if __name__ == '__main__':
#     print(ExcelProcess().write_back(r"F:\workspace_02\unittest_practical\practical_01\test_data\test_data.xlsx","login",2,"pass"))